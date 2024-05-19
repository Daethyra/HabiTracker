import os
import argparse
from openpyxl import load_workbook
from datetime import datetime, timedelta
from utilities import HabiTracker, DatabaseError


def parse_smokes(status: str) -> int:
    if status == "✅":
        return 0
    elif status == "🤘🏼":
        return 1
    elif status == "🚫":
        return 2
    else:
        return status.count("❌") + 2


def ingest_smokes_data(
    filepath: str, habit_name: str, db_name: str = "habitrack.db"
) -> None:
    # Initialize the HabiTracker
    tracker = HabiTracker(db_name)
    tracker.initialize_database()
    tracker.create_necessary_tables()

    # Ensure the habit exists
    try:
        tracker.create_habit(habit_name, "A habit being tracked")
    except DatabaseError as e:
        # If the habit already exists, we can ignore this error
        if "already exists" not in str(e):
            raise

    # Load the workbook and select the active sheet
    wb = load_workbook(filepath)
    sheet = wb.active

    # Iterate over each row in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        date_str = row[0]
        status = row[5]  # Column F contains the status

        # Parse the status to get the number of smokes
        smokes = parse_smokes(status)

        # Check if date_str is already a datetime object
        if isinstance(date_str, datetime):
            date = date_str
        else:
            # Convert date string to datetime object
            date = datetime.strptime(date_str, "%Y-%m-%d")

        # Create entries for each smoke
        for _ in range(smokes):
            # Set the timestamp to 12PM noon
            entry_timestamp = datetime.combine(date, datetime.min.time()) + timedelta(
                hours=12
            )
            tracker.record_habit_entry(habit_name, entry_timestamp)

    print(f"Data ingestion complete for {habit_name} from {filepath}.")


def main():
    parser = argparse.ArgumentParser(
        description="Ingest smoke data from spreadsheets into the habitracker database."
    )
    parser.add_argument("path", nargs="?", help="Directory path or spreadsheet file")
    parser.add_argument("habit_name", nargs="?", help="Name of the habit to track")
    args = parser.parse_args()

    # Prompt for path if not provided
    if not args.path:
        args.path = input("Please enter the directory path or spreadsheet file: ")

    # Prompt for habit name if not provided
    if not args.habit_name:
        args.habit_name = input("Please enter the name of the habit: ")

    # Check if the provided path is a directory or a file
    if os.path.isdir(args.path):
        # Iterate over all 'xlsx' files in the specified directory
        for filename in os.listdir(args.path):
            if filename.endswith(".xlsx"):
                filepath = os.path.join(args.path, filename)
                ingest_smokes_data(filepath, args.habit_name)
    elif os.path.isfile(args.path) and args.path.endswith(".xlsx"):
        # Ingest the single spreadsheet file
        ingest_smokes_data(args.path, args.habit_name)
    else:
        print(
            "Invalid path. Please provide a valid directory path or spreadsheet file."
        )


if __name__ == "__main__":
    main()
